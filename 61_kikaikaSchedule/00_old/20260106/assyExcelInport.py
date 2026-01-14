from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel
from datetime import datetime, timedelta, date
import pandas as pd
import os
import re
import random

# =========================
# 設定
# =========================
FOLDER_PACK = r'\\tiss-ntsrv\TISS-PCS\第三組立課日程（包装機)'
FOLDER_4TH  = r'\\tiss-ntsrv\TISS-PCS\第四組立課日程（フード・ロジ・DRV・DEL)'

WEB_XLSM = r'C:\Users\di2413\OneDrive - DIGIWORLD Cloud\37_WEB日程表\web出荷日程表.xlsm'

# 出力先
combined_file_path = r'\\tiss-ntsrv\TISS-PCS\第三組立課日程（包装機)\Combined_Assembly_Schedule.xlsx'

# 今日〜4週間
today = datetime.now().date()
four_weeks_later = today + timedelta(weeks=4)

# =========================
# 共通関数
# =========================
def normalize(s: str) -> str:
    """検索用に正規化（全角スペース→半角、連続空白まとめ、大小無視）"""
    s = s.replace("\u3000", " ")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s

def parse_date_any(v):
    """openpyxlのセル値からdateへ（datetime/date/文字列/Excelシリアル）"""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, (int, float)):
        # Excelシリアルの可能性
        try:
            return from_excel(v).date()
        except Exception:
            return None
    if isinstance(v, str):
        s = v.strip()
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None
    return None

def find_header_row(sheet, search_terms):
    """ヘッダ行を探す（行のどこかに search_terms のどれかが含まれる行）"""
    for idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row and any(term in row for term in search_terms):
            return idx, row
    return None, None

def count_rows_in_range(file_path: str, start_date: date, end_date: date) -> int:
    """
    指定期間内の「組立開始」行数を数える
    - ヘッダ見つかったシートだけ対象
    - 日付パースできたものだけカウント
    """
    wb = load_workbook(filename=file_path, data_only=True, read_only=True)
    file_name = os.path.basename(file_path)

    total = 0

    for sheet_name in wb.sheetnames:
        sh = wb[sheet_name]

        # まずは通常（PLU / 組立開始）
        header_idx, header_row = find_header_row(sh, ["PLU", "機種"])
        if not header_row:
            continue

        # 列特定（web以外は基本PLU/組立開始）
        try:
            if file_name == 'web出荷日程表.xlsm':
                plu_col = header_row.index("機種")
                start_col = header_row.index("組立開始日")
            else:
                # 一部表記ゆれに備えて「組立開始日」も試す
                plu_col = header_row.index("PLU") if "PLU" in header_row else header_row.index("機種")
                if "組立開始" in header_row:
                    start_col = header_row.index("組立開始")
                elif "組立開始日" in header_row:
                    start_col = header_row.index("組立開始日")
                else:
                    continue
        except ValueError:
            continue

        for row in sh.iter_rows(min_row=header_idx + 1, values_only=True):
            if not row or start_col >= len(row):
                continue
            d = parse_date_any(row[start_col])
            if d and start_date <= d <= end_date:
                total += 1

    try:
        wb.close()
    except Exception:
        pass

    return total

def pick_best_file_by_period(folder: str, must_tokens, start_date: date, end_date: date,
                            ext=(".xlsx", ".xlsm")) -> str:
    """
    1) folder内で must_tokens を全て含むファイル候補を拾う
    2) 指定期間内の行数(count)が最大のファイルを採用
    3) 同点なら更新日時(mtime)が新しい方を採用（タイブレーク）
    """
    if not os.path.isdir(folder):
        raise FileNotFoundError(f"フォルダが見つかりません: {folder}")

    must_tokens_n = [normalize(t) for t in must_tokens]

    candidates = []
    for name in os.listdir(folder):
        if not name.lower().endswith(ext):
            continue
        n = normalize(name)
        if all(t in n for t in must_tokens_n):
            candidates.append(os.path.join(folder, name))

    if not candidates:
        raise FileNotFoundError(f"候補ゼロ: folder={folder}, must={must_tokens}")

    best = None
    best_count = -1
    best_mtime = -1

    for path in candidates:
        try:
            c = count_rows_in_range(path, start_date, end_date)
        except Exception:
            # 壊れてる/開けないファイルはスキップ
            continue

        mtime = os.path.getmtime(path)
        if (c > best_count) or (c == best_count and mtime > best_mtime):
            best = path
            best_count = c
            best_mtime = mtime

    if not best:
        raise FileNotFoundError(f"候補はあるが、どれも読めませんでした: {candidates}")

    return best

# =========================
# ここで “揺れるパス” を自動解決
# =========================
aw_path = pick_best_file_by_period(
    folder=FOLDER_PACK,
    must_tokens=["AW", "組立予定"],
    start_date=today,
    end_date=four_weeks_later,
    ext=(".xlsx",)
)

cp_path = pick_best_file_by_period(
    folder=FOLDER_PACK,
    must_tokens=["CP", "組立予定"],
    start_date=today,
    end_date=four_weeks_later,
    ext=(".xlsx",)
)

fx_path = pick_best_file_by_period(
    folder=FOLDER_PACK,
    must_tokens=["FX", "組立予定"],
    start_date=today,
    end_date=four_weeks_later,
    ext=(".xlsx",)
)

hc_path = pick_best_file_by_period(
    folder=FOLDER_4TH,
    must_tokens=["HC", "組立予定"],
    start_date=today,
    end_date=four_weeks_later,
    ext=(".xlsx",)
)

# 実際に処理するファイル一覧
file_paths = [
    aw_path,
    cp_path,
    fx_path,
    hc_path,
    WEB_XLSM
]

print("採用ファイル:")
for p in file_paths:
    try:
        print(" -", p, " / mtime=", datetime.fromtimestamp(os.path.getmtime(p)))
    except Exception:
        print(" -", p)

# =========================
# 既存ロジック（ほぼそのまま）
# =========================
all_data = []

for file_path in file_paths:
    workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
    file_name = os.path.basename(file_path)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        header_idx, header_row = find_header_row(sheet, ["PLU", "機種"])
        if not header_row:
            continue

        try:
            if file_name == 'web出荷日程表.xlsm':
                plu_col_index = header_row.index("機種")
                assembly_start_col_index = header_row.index("組立開始日")
            else:
                plu_col_index = header_row.index("PLU")
                if "組立開始" in header_row:
                    assembly_start_col_index = header_row.index("組立開始")
                elif "組立開始日" in header_row:
                    assembly_start_col_index = header_row.index("組立開始日")
                else:
                    continue
        except ValueError:
            continue

        for row in sheet.iter_rows(min_row=header_idx + 1, values_only=True):
            if not row or assembly_start_col_index >= len(row) or plu_col_index >= len(row):
                continue

            assembly_start_date = parse_date_any(row[assembly_start_col_index])
            plu_cell = row[plu_col_index]

            if not assembly_start_date:
                continue

            if today <= assembly_start_date <= four_weeks_later:

                if not plu_cell:
                    # AW
                    if 'AW' in file_name:
                        plu_cell = '35502' if random.random() <= 0.6 else '33731'

                    # CP
                    elif 'CP' in file_name:
                        if sheet_name == 'SRX':
                            plu_cell = '40801'
                        else:
                            plu_cell = '35936' if random.random() <= 0.7 else '34921'

                    # FX
                    elif 'FX' in file_name:
                        plu_cell = '34504' if random.random() <= 0.6 else '35449'

                    # HC HI LIS
                    elif 'HC' in file_name:
                        plu_cell = '32923'

                all_data.append((file_name, sheet_name, str(plu_cell), assembly_start_date))

    try:
        workbook.close()
    except Exception:
        pass

all_df = pd.DataFrame(all_data, columns=['Excel File', 'Sheet Name', 'PLU', 'Assembly Start Date'])
all_df.to_excel(combined_file_path, index=False)

print('完了')
